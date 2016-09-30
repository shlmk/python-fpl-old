# -*- coding: utf-8 -*-
"""
Created on Thu Sep 29 18:05:04 2016

@author: MMS
"""

from bs4 import BeautifulSoup
import requests
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import ast
from functools import reduce

#COLORS
RED = 'FF0000'
GREEN = '00B050'
TAN = 'F8CBAD'
YELLOW = 'FFFF00'
ORANGE = 'FFC000'

wb = load_workbook('Premier-League_2016-2017_WeekByWeek.xlsx')
sh_sch = wb['Schedule']
sh_diff = wb['Difficulty']

def color_spreadsheet(week_num, team_h, h_score, team_a, a_score):
    
    #Check game result
    if h_score == a_score: 
        home_color = PatternFill("solid", fgColor=TAN)
        away_color = PatternFill("solid", fgColor=TAN)
    elif h_score > a_score:
        home_color = PatternFill("solid", fgColor=GREEN)
        away_color = PatternFill("solid", fgColor=RED)
    else:
        home_color = PatternFill("solid", fgColor=RED)
        away_color = PatternFill("solid", fgColor=GREEN)

    sh_sch.cell(row=team_h + 1, column=week_num + 1).fill = home_color
    sh_sch.cell(row=team_a + 1, column=week_num + 1).fill = away_color
    sh_diff.cell(row=team_h + 1, column=week_num + 1).fill = home_color
    sh_diff.cell(row=team_a + 1, column=week_num + 1).fill = away_color
    
    #Check actual scores (home score)
    if h_score == 0:
        home_off_color = PatternFill("solid", fgColor=RED)
        away_def_color = PatternFill("solid", fgColor=GREEN)
    elif h_score == 1:
        home_off_color = PatternFill("solid", fgColor=ORANGE)
        away_def_color = PatternFill("solid", fgColor=YELLOW)
    elif h_score == 2:
        home_off_color = PatternFill("solid", fgColor=YELLOW)
        away_def_color = PatternFill("solid", fgColor=ORANGE)
    else:
        home_off_color = PatternFill("solid", fgColor=GREEN)
        away_def_color = PatternFill("solid", fgColor=RED)
    
    sh_sch.cell(row=team_h + 45, column=week_num + 1).fill = home_off_color
    sh_sch.cell(row=team_a + 23, column=week_num + 1).fill = away_def_color
    sh_diff.cell(row=team_h + 45, column=week_num + 1).fill = home_off_color
    sh_diff.cell(row=team_a + 23, column=week_num + 1).fill = away_def_color
   
   #Check actual scores (away score)
    if a_score == 0:
        home_off_color = PatternFill("solid", fgColor=GREEN)
        away_def_color = PatternFill("solid", fgColor=RED)
    elif a_score == 1:
        home_off_color = PatternFill("solid", fgColor=YELLOW)
        away_def_color = PatternFill("solid", fgColor=ORANGE)
    elif a_score == 2:
        home_off_color = PatternFill("solid", fgColor=ORANGE)
        away_def_color = PatternFill("solid", fgColor=YELLOW)
    else:
        home_off_color = PatternFill("solid", fgColor=RED)
        away_def_color = PatternFill("solid", fgColor=GREEN)
    
    sh_sch.cell(row=team_h + 23, column=week_num + 1).fill = home_off_color
    sh_sch.cell(row=team_a + 45, column=week_num + 1).fill = away_def_color
    sh_diff.cell(row=team_h + 23, column=week_num + 1).fill = home_off_color
    sh_diff.cell(row=team_a + 45, column=week_num + 1).fill = away_def_color
       
       
def process_week(week_html, week_num):
    week = week_html.text
    
    #From: https://stackoverflow.com/questions/6116978/python-replace-multiple-strings
    repls = ('true','True'), ('false', 'False'), ('null', '[]')
    week = reduce(lambda a, kv: a.replace(*kv), repls, week)
    week = ast.literal_eval(week)
    
    for game in week: 
        color_spreadsheet(week_num, game['team_h'], game['team_h_score'], 
                          game['team_a'], game['team_a_score'])
                          
    wb.save('Premier-League_2016-2017_WeekByWeek.xlsx')
            
#NOTE: weeks must be a list format
def record_data(weeks):
    base_url = 'https://fantasy.premierleague.com/drf/fixtures/?event='

    if type(weeks) is not list:
        raise TypeError('Weeks must be of type list')
    else:
        for week_num in weeks:
            if type(week_num) is not int:
                raise TypeError('Week must be an int value')
            elif week_num < 1 or week_num > 38:
                raise IndexError('Week must be between 1 and 38')
            else: 
                url = base_url + str(week_num)
                
            r = requests.get(url)
            soup = BeautifulSoup(r.text, 'lxml')
            
            process_week(soup, week_num)
            
record_data(list(range(1,7)))


          